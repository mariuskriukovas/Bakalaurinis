import numpy as np
import pandas as pd
from openpyxl import load_workbook

NAME = 'C:\\Users\\Marius\\PycharmProjects\\Kursinis\\git\\Bakalaurinis\\results\\database.xlsx'


def write_to_excel(df_arr, df_names, file_name = NAME):
    with pd.ExcelFile(file_name) as writer:
        for i, df in enumerate(df_arr):
            sheet_name = f'Sheet_{df_names[i]}'
            df.to_excel(writer, sheet_name=sheet_name)
    writer.save()
    writer.close()


def append_excel_sheets(df_arr, df_names):
    book = load_workbook(NAME)
    writer = pd.ExcelWriter(NAME, engine='openpyxl')
    writer.book = book
    for i, df in enumerate(df_arr):
        sheet_name = f'Sheet_{df_names[i]}'
        df.to_excel(writer, sheet_name=sheet_name)
    writer.save()
    writer.close()


def get_excel_sheets(names) -> dict:
    df_dic = {}
    for n in names:
        df_dic[n] = pd.read_excel(NAME, sheet_name=n).drop(columns=['Unnamed: 0'])
    return df_dic


def get_excel_sheet_names():
    book = load_workbook(NAME)
    return book.get_sheet_names()

def print_excel_sheets():
    for e in get_excel_sheet_names():
        print(e)